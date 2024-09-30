import csv
from openpyxl import load_workbook
import json


class Gelcount:
    def __init__(self,path):
        self.data = self.__read_data(path)

    def get_data(self):
        return json.dumps(self.data)

    def __read_data(self,path):
        data = list()
        if path.endswith(".csv"):
            data = self.__read_csv_file(path)
        elif path.endswith(".xlsx"):
            data = self.__read_xlsx_file(path)
        final_data = []

        keys = []
        units = []

        for key in data[0]:
            unit = ''
            if '(' in key:
                start = key.find('(')
                end = key.find(')')
                unit = key[start + 1:end]
                key = key[:start - 1] + key[end + 1:]
            keys.append(key)
            units.append(unit)

        for line in data[1:]:
            result = {}
            for i, key in enumerate(keys):
                result[key] = {'unit': units[i], 'value': str(line[i])}
            final_data.append(result)

        return final_data
    # def read_data(path):
    #     data = list()
    #     if path.endswith(".csv"):
    #         data = read_csv_file(path)
    #     elif path.endswith(".xlsx"):
    #         data = read_xlsx_file(path)
    #     final_data = []
    #     val, result = {}, {}
    #
    #     for key in data[0]:
    #         unit = ''
    #         if '(' in key:
    #             start = key.find('(')
    #             end = key.find(')')
    #             unit = key[start + 1:end]
    #             key = key[:start - 1] + key[end + 1:]
    #         val[key] = {'unit': unit, 'value': ''}
    #
    #     for line in data[1:]:
    #         result = copy.deepcopy(val)  # Create a deep copy of val
    #         count = 0
    #         for key in result.keys():
    #             result[key]['value'] = str(line[count])
    #             count += 1
    #         final_data.append(result)
    #     return final_data

    def __read_csv_file(self,file_path):
        with open(file_path, 'r') as f:
            reader = csv.reader(f)
            data = list(reader)
        return data

    def __read_xlsx_file(self,file_path):
        wb = load_workbook(file_path)
        sheet_names = wb.sheetnames
        # Select a sheet
        sheet = wb[sheet_names[0]]
        data = []
        # Iterate through each row in the sheet
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))
        return data

