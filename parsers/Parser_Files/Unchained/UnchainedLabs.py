import openpyxl
import json
import TableFromExcel
from datetime import datetime

#Parser for Lunatic Instrument Files
class Lunatic:
    # Constructor method to initialize the file_path
    def __init__(self, file_path):
        self.file_path = file_path

    #Fetch the Info Part from the File
    def getInfo(self):
        filename= self.file_path
        try:
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.active

            data_dict = {}

            start_reading = False

            for row in sheet.iter_rows(values_only=True):
                if not start_reading:
                    if row[0] is not None:
                        start_reading = True
                if start_reading:
                    if row[0] is None:  
                        break
                    if isinstance(row[1], datetime):
                        data_dict[row[0]] = str(row[1])
                    else:
                        data_dict[row[0]] = row[1]

            keys = list(data_dict.keys())
            new_dict = {key: data_dict[key] for key in keys[1:]}

            return json.dumps(new_dict, indent=1)
        except Exception as e:
            return str(e)


    def __find_table_index(self):
        filename = self.file_path
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active

        table_index = None

        for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row[0] == "Table":
                table_index = index
                break

        return table_index
    
    #Fetch the Plate Informations from the File
    def getPlateInfo(self):
        filename=self.file_path
        try:
            workbook = openpyxl.load_workbook(filename)
            table_index = self.__find_table_index()
            table_start_index = table_index+1
            first_sheet_name = workbook.sheetnames[0]

            plate_data_list = TableFromExcel.fetch_data_from_excel(filename, first_sheet_name, table_start_index)

            return json.dumps(plate_data_list)
        except Exception as e:
            return str(e)


