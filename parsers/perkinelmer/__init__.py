import os
import pandas as pd
import json
from openpyxl import load_workbook
import perkinelmer.TableFromExcel


#Parser for Dropsense_96 Instrument
class Dropsense96:
    # Constructor method to initialize the file_path
    def __init__(self, file_path):
        self.file_path = file_path

    # Function to get the Basic Informations
    def get_info(self):
        try:
            wb = load_workbook(filename=self.file_path)
            sheet = wb['cDrop Report']

            l = []

            for row in sheet.iter_rows():
                for cell in row:
                    cell_value = str(cell.value)
                    l.append(cell_value)

            l = [i for i in l if i != 'None']

            extracted_data_dict = {}

            extracted_data_dict_keys = [
                "Date",
                "Test Performed By",
                "Instrument",
                "Software Version",
                "DropQuant software version",
                "Experiment Name",
                "Application",
                "Blanks",
                "DropPlate/Slide type",
                "Nr of DropPlates"
            ]

            extracted_data_dict_keys_in_excel_list = [
                "Date",
                "Test performed by",
                "Instrument",
                "Software version",
                "DropQuant software version",
                "Experiment name",
                "Application",
                "Blanks",
                "DropPlate/Slide type",
                "Nr of DropPlates"
            ]

            for dict_key, value_in_list in zip(extracted_data_dict_keys, extracted_data_dict_keys_in_excel_list):
                if value_in_list in l:
                    extracted_data_dict[dict_key] = l[l.index(value_in_list) + 1].strip()
                else:
                    continue

            return json.dumps({'info':extracted_data_dict})
        except Exception as e:
            return str(e)

    # Function to get the number of plates
    def get_plateNum(self):
        try:
            dropplate_data_list = TableFromExcel.fetch_data_from_excel(self.file_path, 'cDrop Report', 17)
            dropplate_df = pd.DataFrame(dropplate_data_list)
            unique_plates = set(dropplate_df['DropPlate ID'].tolist())

            return json.dumps({'plate count': len(unique_plates)})
        except Exception as e:
            return str(e)

    # Function to get the Unique Plate IDs
    def get_plateId(self):
        try:
            dropplate_data_list = TableFromExcel.fetch_data_from_excel(self.file_path, 'cDrop Report', 17)
            dropplate_df = pd.DataFrame(dropplate_data_list)
            unique_plates = list(set(dropplate_df['DropPlate ID'].tolist()))

            return json.dumps({'ID':unique_plates})
        except Exception as e:
            return str(e)

    # Function to get the Plate Details by providing Plate ID
    def get_plateInfo(self, plateId):
        try:
            dropplate_data_list = TableFromExcel.fetch_data_from_excel(self.file_path, 'cDrop Report', 17)
            dropplate_df = pd.DataFrame(dropplate_data_list)

            final_plate_dict = {}
            final_plate_dict['Plate Id'] = plateId
            each_dropplate_df = dropplate_df[dropplate_df['DropPlate ID'] == plateId]
            each_dropplate_df = each_dropplate_df.loc[:,
                                ['DropPlate\nPosition', 'Sample name', 'Pump', 'A280 Concentration\n(mg/ml)', 'E1%',
                                 'Background (A280)', 'A280', 'A260/A280']]
            each_dropplate_df = each_dropplate_df.rename(columns={'DropPlate\nPosition': 'Position',
                                                                  'A280 Concentration\n(mg/ml)': 'A280 Concentration (mg/ml)'})
            each_dropplate_list = each_dropplate_df.to_dict(orient='records')
            final_plate_dict['Wells'] = each_dropplate_list

            return json.dumps({'plate info': final_plate_dict})
        except Exception as e:
            return str(e)

    # Function to get All Plate Details
    def get_allPlateInfo(self):
        try:
            dropplate_data_list = TableFromExcel.fetch_data_from_excel(self.file_path, 'cDrop Report', 17)
            dropplate_df = pd.DataFrame(dropplate_data_list)
            unique_plates = list(set(dropplate_df['DropPlate ID'].tolist()))
            all_plate_list = []
            for plate in unique_plates:
                each_plate = {}
                each_plate['Plate ID'] = plate
                each_dropplate_df = dropplate_df[dropplate_df['DropPlate ID'] == plate]
                each_dropplate_df = each_dropplate_df.loc[:,
                                    ['DropPlate\nPosition', 'Sample name', 'Pump', 'A280 Concentration\n(mg/ml)', 'E1%',
                                     'Background (A280)', 'A280', 'A260/A280']]
                each_dropplate_df = each_dropplate_df.rename(columns={'DropPlate\nPosition': 'Position',
                                                                      'A280 Concentration\n(mg/ml)': 'A280 Concentration (mg/ml)'})
                each_dropplate_list = each_dropplate_df.to_dict(orient='records')
                each_plate['Wells'] = each_dropplate_list

                all_plate_list.append(each_plate)

            return json.dumps({'plates':all_plate_list})
        except Exception as e:
            return str(e)


#Parser for EnvisionXcite2105 Instrument
class Envisionxcite:
    # Constructor method to initialize the file_path
    def __init__(self, file_path):
        self.file_path = file_path

    #Function to fetch Plate Information
    def get_plateInfo(self):
        try:
            with open(self.file_path, "r") as doc:
                data_list = doc.read().split('\n')
                data_details_list = []

                for data_value_list in data_list:
                    data_details_list.extend(data_value_list.split('\t'))

                plate_information_input_list = data_details_list[:data_details_list.index('Background information')]

                plate_information_dict_keys = plate_information_input_list[1:12]
                plate_information_list = []
                plate_information_dict = {}

                for plate_information_key in plate_information_dict_keys:
                    plate_information_dict[plate_information_key.strip()] = "NIL" if plate_information_input_list[
                                                                                         plate_information_input_list.index(
                                                                                             plate_information_key) + 12].strip() == "" else \
                    plate_information_input_list[plate_information_input_list.index(plate_information_key) + 12].strip()
                plate_information_list.append(plate_information_dict)

            return json.dumps({'info':plate_information_list})
        except Exception as e:
            return str(e)

    #Function to fetch Background Information
    def get_backgound(self):
        try:
            with open(self.file_path, "r") as doc:
                data_list = doc.read().split('\n')
                data_details_list = []

                for data_value_list in data_list:
                    data_details_list.extend(data_value_list.split('\t'))

                background_information_input_list = data_details_list[data_details_list.index(
                    'Background information') + 1:data_details_list.index('Background information') + 25]

                background_information_dict_keys = background_information_input_list[:7]

                background_information_list = []

                number_of_background_rows = len(background_information_input_list) // 8 - 1

                for row in range(1, number_of_background_rows + 1):
                    background_information_dict = {}
                    for background_information_key in background_information_dict_keys:
                        background_information_dict[background_information_key.strip()] = "NIL" if \
                        background_information_input_list[background_information_input_list.index(
                            background_information_key) + 8 * row].strip() == "" else background_information_input_list[
                            background_information_input_list.index(background_information_key) + 8 * row].strip()
                    background_information_list.append(background_information_dict)

            return json.dumps({'background':background_information_list})
        except Exception as e:
            return str(e)

    #Function to fetch All Plate Details
    def get_allPlateInfo(self):
        try:
            with open(self.file_path, "r") as doc:
                data_list = doc.read().split('\n')
                data_details_list = []

                for data_value_list in data_list:
                    data_details_list.extend(data_value_list.split('\t'))

                width_list = data_details_list[
                             data_details_list.index('Background information') + 25:data_details_list.index(
                                 'Background information') + 35]
                width_list = [w for w in width_list if w != ""]

                values_input_list = data_details_list[
                                    data_details_list.index('CalcResultI') - 3:data_details_list.index(
                                        'Basic assay information ') - 1]

                values_list = []

                number_of_value_rows = len(values_input_list) // 13 - 1

                for row in range(1, number_of_value_rows + 1):
                    values_dict = {}

                    values_dict['Plate'] = 'NIL' if values_input_list[
                                                        values_input_list.index('Plate') + 13 * row].strip() == "" else \
                    values_input_list[values_input_list.index('Plate') + 13 * row].strip()
                    values_dict['Barcode'] = 'NIL' if values_input_list[values_input_list.index(
                        'Barcode') + 13 * row].strip() == "" else values_input_list[
                        values_input_list.index('Barcode') + 13 * row].strip()
                    values_dict['Well'] = 'NIL' if values_input_list[
                                                       values_input_list.index('Well') + 13 * row].strip() == "" else \
                    values_input_list[values_input_list.index('Well') + 13 * row].strip()

                    first_inner_child_list = []

                    for width_row in range(0, len(width_list)):
                        second_inner_child_dict = {}

                        second_inner_child_dict['width'] = width_list[width_row]

                        second_inner_child_dict = {}

                        second_inner_child_dict['CalcResultI'] = 'NIL' if values_input_list[values_input_list.index(
                            'CalcResultI') + 13 * row + width_row * 4] == "" else values_input_list[
                            values_input_list.index('CalcResultI') + 13 * row + width_row * 4]
                        second_inner_child_dict['Signal'] = 'NIL' if values_input_list[values_input_list.index(
                            'Signal') + 13 * row + width_row * 4] == "" else values_input_list[
                            values_input_list.index('Signal') + 13 * row + width_row * 4]
                        second_inner_child_dict['Flashes/Time'] = 'NIL' if values_input_list[values_input_list.index(
                            'Flashes/Time') + 13 * row + width_row * 4] == "" else values_input_list[
                            values_input_list.index('Flashes/Time') + 13 * row + width_row * 4]
                        second_inner_child_dict['MeasTime'] = 'NIL' if values_input_list[values_input_list.index(
                            'MeasTime') + 13 * row + width_row * 4] == "" else values_input_list[
                            values_input_list.index('MeasTime') + 13 * row + width_row * 4]

                        first_inner_child_list.append(second_inner_child_dict)

                    values_dict['child'] = first_inner_child_list

                    values_list.append(values_dict)

            return json.dumps({'plates info':values_list})
        except Exception as e:
            return str(e)

    #Function to fetch Basic Assay Information
    def get_basicAssayInfo(self):
        try:
            with open(self.file_path, "r") as doc:
                data_list = doc.read().split('\n')
                data_details_list = []

                for data_value_list in data_list:
                    data_details_list.extend(data_value_list.split('\t'))

                basic_assay_informations_input_list = data_details_list[data_details_list.index(
                    'Basic assay information ') + 1:data_details_list.index('Protocol information')]

                basic_assay_informations_input_list = [i for i in basic_assay_informations_input_list if i != ""]

                basic_assay_informations_list = []

                basic_assay_informations_dict = {}

                basic_assay_informations_dict_keys = [bai for bai in basic_assay_informations_input_list if
                                                      basic_assay_informations_input_list.index(bai) % 2 == 0]

                for basic_assay_informations_key in basic_assay_informations_dict_keys:
                    basic_assay_informations_dict[basic_assay_informations_key] = 'NIL' if \
                    basic_assay_informations_input_list[
                        basic_assay_informations_input_list.index(basic_assay_informations_key) + 1] == "" else \
                    basic_assay_informations_input_list[
                        basic_assay_informations_input_list.index(basic_assay_informations_key) + 1]

                basic_assay_informations_list.append(basic_assay_informations_dict)

            return json.dumps({'assay info' : basic_assay_informations_list})
        except Exception as e:
            return str(e)

    #Function to fetch Protocol Information
    def get_protocolInfo(self):
        try:
            with open(self.file_path, "r") as doc:
                data_list = doc.read().split('\n')
                data_details_list = []

                for data_value_list in data_list:
                    data_details_list.extend(data_value_list.split('\t'))

                protocol_informations_input_list = data_details_list[
                                                   data_details_list.index('Protocol:') + 1:data_details_list.index(
                                                       'Plate type:')]

                protocol_informations_input_index_list = [i for i in protocol_informations_input_list if i != ""]

                protocol_informations_list = []

                protocol_informations_dict = {}

                protocol_informations_dict_keys = [pi for pi in protocol_informations_input_index_list if
                                                   protocol_informations_input_index_list.index(pi) % 2 == 0]

                for protocol_informations_key in protocol_informations_dict_keys:
                    protocol_informations_dict[protocol_informations_key] = 'NIL' if protocol_informations_input_list[
                                                                                         protocol_informations_input_list.index(
                                                                                             protocol_informations_key) + 4].strip() == "" else \
                    protocol_informations_input_list[
                        protocol_informations_input_list.index(protocol_informations_key) + 4].strip()

                protocol_informations_list.append(protocol_informations_dict)

                return json.dumps({'protocol':protocol_informations_list})
        except Exception as e:
            return str(e)

    #Function to fetch Plate Type Information
    def get_plateTypeInfo(self):
        try:
            with open(self.file_path, "r") as doc:
                data_list = doc.read().split('\n')
                data_details_list = []

                for data_value_list in data_list:
                    data_details_list.extend(data_value_list.split('\t'))

                plate_type_input_list = data_details_list[
                                        data_details_list.index('Plate type:') + 1:data_details_list.index(
                                            'Coordinates of corners:')]

                plate_type_input_index_list = [i for i in plate_type_input_list if i != ""]

                plate_type_list = []

                plate_type_dict = {}

                plate_type_dict_keys = [plt for plt in plate_type_input_index_list if
                                        plate_type_input_index_list.index(plt) % 2 == 0]

                for plate_type_key in plate_type_dict_keys:
                    plate_type_dict[plate_type_key] = 'NIL' if plate_type_input_list[plate_type_input_list.index(
                        plate_type_key) + 4].strip() == "" else plate_type_input_list[
                        plate_type_input_list.index(plate_type_key) + 4].strip()

                plate_type_list.append(plate_type_dict)

            return json.dumps({'plate type':plate_type_list})
        except Exception as e:
            return str(e)

    #Function to fetch Auto Export Parameters Information
    def get_autoExportParaInfo(self):
        try:
            with open(self.file_path, "r") as doc:
                data_list = doc.read().split('\n')
                data_details_list = []

                for data_value_list in data_list:
                    data_details_list.extend(data_value_list.split('\t'))

                auto_export_parameters_input_list = data_details_list[data_details_list.index(
                    'Auto export parameters:') + 1:data_details_list.index('Operations:')]

                auto_export_parameters_input_index_list = [i for i in auto_export_parameters_input_list if i != ""]

                auto_export_parameters_list = []

                auto_export_parameters_dict = {}

                auto_export_parameters_dict_keys = [aep for aep in auto_export_parameters_input_index_list if
                                                    auto_export_parameters_input_index_list.index(aep) % 2 == 0]

                for auto_export_parameters_key in auto_export_parameters_dict_keys:
                    auto_export_parameters_dict[auto_export_parameters_key] = 'NIL' if \
                    auto_export_parameters_input_list[
                        auto_export_parameters_input_list.index(auto_export_parameters_key) + 4].strip() == "" else \
                    auto_export_parameters_input_list[
                        auto_export_parameters_input_list.index(auto_export_parameters_key) + 4].strip()

                auto_export_parameters_list.append(auto_export_parameters_dict)

            return json.dumps({'export parameters':auto_export_parameters_list})
        except Exception as e:
            return str(e)

    #Function to fetch Operations Information
    def get_operationInfo(self):
        try:
            with open(self.file_path, "r") as doc:
                data_list = doc.read().split('\n')
                data_details_list = []

                for data_value_list in data_list:
                    data_details_list.extend(data_value_list.split('\t'))

                operations_input_list = data_details_list[
                                        data_details_list.index('Operations:') + 1:data_details_list.index('Labels:')]

                operations_input_index_list = [i for i in operations_input_list if i != ""]

                operations_list = []

                operations_dict = {}

                operations_dict_keys = ['Plate 1', 'Group 1', 'Measurement', '      Label']

                for operations_key in operations_dict_keys:
                    if operations_key == '      Label':
                        operations_dict[operations_key.strip()] = 'NIL' if operations_input_list[
                                                                               operations_input_list.index(
                                                                                   operations_key) + 4].strip() == "" else \
                        operations_input_list[operations_input_list.index(operations_key) + 4].strip()
                    else:
                        operations_dict[operations_key.strip()] = 'NIL'
                operations_list.append(operations_dict)

            return json.dumps({'operation':operations_list})
        except Exception as e:
            return str(e)

    #Function to fetch Label Information
    def get_labelInfo(self):
        try:
            with open(self.file_path, "r") as doc:
                data_list = doc.read().split('\n')
                data_details_list = []

                for data_value_list in data_list:
                    data_details_list.extend(data_value_list.split('\t'))

                labels_input_list = data_details_list[
                                    data_details_list.index('Labels:') + 1:data_details_list.index('Filters:')]

                labels_input_index_list = [i for i in labels_input_list if i != ""]

                labels_list = []

                labels_dict = {}

                labels_dict_keys = [lbl for lbl in labels_input_index_list if
                                    labels_input_index_list.index(lbl) % 2 == 0]

                for labels_key in labels_dict_keys:
                    labels_dict[labels_key] = 'NIL' if labels_input_list[
                                                           labels_input_list.index(labels_key) + 4].strip() == "" else \
                    labels_input_list[labels_input_list.index(labels_key) + 4].strip()

                labels_list.append(labels_dict)

            return json.dumps({'label':labels_list})
        except Exception as e:
            return str(e)

    #Function to fetch Filter Information
    def get_filterInfo(self):
        try:
            with open(self.file_path, "r") as doc:
                data_list = doc.read().split('\n')
                data_details_list = []

                for data_value_list in data_list:
                    data_details_list.extend(data_value_list.split('\t'))

                filters_input_list = data_details_list[
                                     data_details_list.index('Filters:') + 1:data_details_list.index('Mirror modules:')]

                number_of_filter_rows = len(filters_input_list) // 36

                filters_list = []

                for filter_row in range(1, number_of_filter_rows + 1):
                    filters_row_dict = {}
                    filters_row_input_list = filters_input_list[(filter_row - 1) * 36:filter_row * 36]

                    filters_row_input_index_list = [i for i in filters_row_input_list if i != ""]
                    filters_row_dict_keys = [filt for filt in filters_row_input_index_list if
                                             filters_row_input_index_list.index(filt) % 2 == 0]

                    for filters_row_key in filters_row_dict_keys:
                        filters_row_dict[filters_row_key] = 'NIL' if filters_row_input_list[
                                                                         filters_row_input_list.index(
                                                                             filters_row_key) + 4].strip() == "" else \
                        filters_row_input_list[filters_row_input_list.index(filters_row_key) + 4].strip()

                    filters_list.append(filters_row_dict)

            return json.dumps({'filter':filters_list})
        except Exception as e:
            return str(e)

    #Function to fetch Mirror Module Information
    def get_mirrorModuleInfo(self):
        try:
            with open(self.file_path, "r") as doc:
                data_list = doc.read().split('\n')
                data_details_list = []

                for data_value_list in data_list:
                    data_details_list.extend(data_value_list.split('\t'))
                mirror_modules_input_list = data_details_list[
                                            data_details_list.index('Mirror modules:') + 1:data_details_list.index(
                                                'Instrument:')]

                mirror_modules_input_index_list = [i for i in mirror_modules_input_list if i != ""]

                mirror_modules_list = []

                mirror_modules_dict = {}

                mirror_modules_dict_keys = [mm for mm in mirror_modules_input_index_list if
                                            mirror_modules_input_index_list.index(mm) % 2 == 0]

                for mirror_modules_key in mirror_modules_dict_keys:
                    mirror_modules_dict[mirror_modules_key] = 'NIL' if mirror_modules_input_list[
                                                                           mirror_modules_input_list.index(
                                                                               mirror_modules_key) + 4].strip() == "" else \
                    mirror_modules_input_list[mirror_modules_input_list.index(mirror_modules_key) + 4].strip()

                mirror_modules_list.append(mirror_modules_dict)

            return json.dumps({'mirror module':mirror_modules_list})
        except Exception as e:
            return str(e)

    #Function to fetch Instrument Information
    def get_instrumentInfo(self):
        try:
            with open(self.file_path, "r") as doc:
                data_list = doc.read().split('\n')
                data_details_list = []

                for data_value_list in data_list:
                    data_details_list.extend(data_value_list.split('\t'))

                instrument_input_list = data_details_list[
                                        data_details_list.index('Instrument:') + 1:data_details_list.index(
                                            'Normalization:')]

                instrument_input_index_list = [i for i in instrument_input_list if i != ""]

                instrument_list = []

                instrument_dict = {}

                instrument_dict_keys = [ins for ins in instrument_input_index_list if
                                        instrument_input_index_list.index(ins) % 2 == 0]

                for instrument_key in instrument_dict_keys:
                    instrument_dict[instrument_key] = 'NIL' if instrument_input_list[instrument_input_list.index(
                        instrument_key) + 4].strip() == "" else instrument_input_list[
                        instrument_input_list.index(instrument_key) + 4].strip()

                instrument_list.append(instrument_dict)

            return json.dumps({'instrument':instrument_list})
        except Exception as e:
            return str(e)
